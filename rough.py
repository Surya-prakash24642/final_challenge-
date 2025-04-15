from openpyxl import load_workbook

# Load workbook and sheet
wb = load_workbook("updated.xlsx")
ws = wb.active

# Loop through each row
for row in ws.iter_rows(min_row=1):
    col1_value = row[0].value

    if col1_value == "Standard Life Versicherung Zweigndl. Dtl. d. Standard Life Int. DAC":
        # Explicitly force string overwrite (removes any object type)
#         row[1].value = None  # clear whatever object was there
#         row[1].data_type = 's'  # force it to be a string cell
#         row[1].value = "https://www.standardlife.de/"
        row[1].value = "https://www.standardlife.de/"
        print(f"Updated URL for {col1_value} to {row[1].value}")

# # Save
# wb.save("updated_file.xlsx")
